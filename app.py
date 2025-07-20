from flask import Flask, render_template, request, redirect
import sqlite3
import os
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)

def init_db():
    if not os.path.exists("database.db"):
        conn = sqlite3.connect("database.db")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                quantity INTEGER NOT NULL
            );
        """)
        conn.close()

@app.route('/')
def index():
    conn = sqlite3.connect("database.db")
    cursor = conn.execute("SELECT * FROM products")
    products = cursor.fetchall()
    conn.close()
    return render_template("index.html", products=products)

@app.route('/add', methods=['GET', 'POST'])
def add_product():
    if request.method == 'POST':
        name = request.form['name']
        quantity = int(request.form['quantity'])
        conn = sqlite3.connect("database.db")
        conn.execute("INSERT INTO products (name, quantity) VALUES (?, ?)", (name, quantity))
        conn.commit()
        conn.close()
        return redirect('/')
    return render_template("add.html")

@app.route('/sell/<int:product_id>')
def sell_product(product_id):
    amount = request.args.get("amount", default=1, type=int)
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT quantity FROM products WHERE id = ?", (product_id,))
    result = cursor.fetchone()
    if result and result[0] >= amount:
        cursor.execute("UPDATE products SET quantity = quantity - ? WHERE id = ?", (amount, product_id))
        conn.commit()
    conn.close()
    return redirect('/')

@app.route('/delete/<int:product_id>')
def delete_product(product_id):
    conn = sqlite3.connect("database.db")
    conn.execute("DELETE FROM products WHERE id = ?", (product_id,))
    conn.commit()
    conn.close()
    return redirect('/')

@app.route('/import_excel', methods=['GET', 'POST'])
def import_excel():
    if request.method == 'POST':
        file = request.files['excel_file']
        filename = secure_filename(file.filename)
        filepath = os.path.join('static', filename)
        file.save(filepath)

        wb = load_workbook(filepath)
        ws = wb.active

        conn = sqlite3.connect("database.db")
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, quantity = row
            if name and quantity:
                conn.execute("INSERT INTO products (name, quantity) VALUES (?, ?)", (name, int(quantity)))
        conn.commit()
        conn.close()
        os.remove(filepath)
        return redirect('/')
    return render_template("import_excel.html")

if __name__ == "__main__":
    init_db()
    app.run(debug=True)
